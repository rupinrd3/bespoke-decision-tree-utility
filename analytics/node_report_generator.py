#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Node Report Generator for Bespoke Utility
Generates detailed reports for decision tree nodes.

"""

import logging
import pandas as pd
import numpy as np
import gc
import time
from typing import Dict, List, Tuple, Optional, Union, Any
from datetime import datetime
from pathlib import Path

from models.decision_tree import BespokeDecisionTree
from models.node import TreeNode

logger = logging.getLogger(__name__)

class NodeReportGenerator:
    """Generates comprehensive node reports from decision tree models"""
    
    def __init__(self):
        """Initialize the node report generator"""
        self.cached_reports = {}
        self.max_depth = 50  # Prevent infinite recursion
        
    def generate_node_report(self, model: BespokeDecisionTree, dataset: pd.DataFrame, 
                           target_variable: str, max_nodes: int = 1000) -> pd.DataFrame:
        """
        Generate comprehensive node report from decision tree model
        
        Args:
            model: Fitted BespokeDecisionTree instance
            dataset: Dataset used for evaluation
            target_variable: Name of target variable column
            max_nodes: Maximum number of nodes to include in report
            
        Returns:
            DataFrame with node report data
            
        Raises:
            ValueError: If inputs are invalid
            RuntimeError: If report generation fails
        """
        try:
            self._validate_inputs(model, dataset, target_variable)
            
            logger.info(f"Generating node report for {len(dataset)} records")
            
            terminal_nodes = self._extract_terminal_nodes(model.root)
            logger.debug(f"Found {len(terminal_nodes)} terminal nodes")
            
            node_number_mapping = self._create_node_number_mapping(model.root)
            
            if not terminal_nodes:
                raise RuntimeError("No terminal nodes found in decision tree")
            
            if len(terminal_nodes) > max_nodes:
                logger.warning(f"Limiting nodes to {max_nodes} out of {len(terminal_nodes)}")
                terminal_nodes = terminal_nodes[:max_nodes]
            
            overall_stats = self._get_tree_level_statistics(model)
            
            node_data = []
            failed_nodes = []
            
            for idx, node in enumerate(terminal_nodes):
                try:
                    node_number = node_number_mapping.get(node.node_id, idx + 1)  # Fallback to sequential if not found
                    node_stats = self._calculate_node_statistics(node, overall_stats, node_number)
                    if node_stats:  # Only add if statistics were calculated successfully
                        node_data.append(node_stats)
                    else:
                        failed_nodes.append(getattr(node, 'node_id', f'node_{idx}'))
                except Exception as e:
                    node_id = getattr(node, 'node_id', f'node_{idx}')
                    logger.warning(f"Failed to calculate statistics for node {node_id}: {e}")
                    failed_nodes.append(node_id)
                    continue
            
            if not node_data:
                raise RuntimeError("No valid node statistics calculated")
            
            if failed_nodes:
                logger.warning(f"Failed to process {len(failed_nodes)} nodes: {failed_nodes}")
            
            self._validate_calculation_results(node_data, overall_stats)
            
            df = pd.DataFrame(node_data)
            df = self._calculate_cumulative_metrics(df)
            
            df = df.sort_values('node_size', ascending=False).reset_index(drop=True)
            
            logger.info(f"Generated node report with {len(df)} nodes")
            return df
            
        except Exception as e:
            logger.error(f"Failed to generate node report: {e}")
            raise RuntimeError(f"Node report generation failed: {str(e)}")
    
    def _validate_inputs(self, model: BespokeDecisionTree, dataset: pd.DataFrame, 
                        target_variable: str) -> None:
        """Comprehensive input validation for node report generation"""
        if not isinstance(model, BespokeDecisionTree):
            raise ValueError("Model must be BespokeDecisionTree instance")
        
        if not hasattr(model, 'is_fitted') or not model.is_fitted:
            raise ValueError("Model must be fitted before generating report")
        
        if dataset is None or dataset.empty:
            raise ValueError("Dataset cannot be empty")
        
        if len(dataset) == 0:
            raise ValueError("Dataset has no rows")
        
        if target_variable not in dataset.columns:
            raise ValueError(f"Target variable '{target_variable}' not found in dataset")
        
        if dataset[target_variable].isna().all():
            raise ValueError("Target variable contains only NaN values")
        
        unique_target_values = dataset[target_variable].dropna().nunique()
        if unique_target_values <= 1:
            raise ValueError("Target variable must have at least 2 unique values for meaningful analysis")
        
        if not hasattr(model, 'root') or model.root is None:
            raise ValueError("Model does not have a valid tree structure")
        
        if hasattr(model, 'feature_names') and model.feature_names:
            model_features = set(model.feature_names)
            dataset_features = set(dataset.columns) - {target_variable}
            missing_features = model_features - dataset_features
            if missing_features:
                raise ValueError(f"Model features not found in dataset: {missing_features}")
        
        target_distribution = dataset[target_variable].value_counts(normalize=True)
        min_class_proportion = target_distribution.min()
        if min_class_proportion < 0.01:  # Less than 1%
            logger.warning(f"Extremely unbalanced target distribution detected. "
                         f"Minimum class proportion: {min_class_proportion:.3f}")
        
        logger.debug(f"Input validation passed: {len(dataset)} records, {unique_target_values} target classes")
    
    def _extract_terminal_nodes(self, root_node: TreeNode) -> List[TreeNode]:
        """
        Safely extract terminal nodes from decision tree
        
        Args:
            root_node: Root node of the decision tree
            
        Returns:
            List of terminal nodes
        """
        if not root_node:
            return []
        
        terminal_nodes = []
        visited = set()
        
        def traverse(node: TreeNode, depth: int = 0) -> None:
            if depth > self.max_depth:
                logger.warning("Maximum tree depth exceeded during traversal")
                return
            
            if id(node) in visited:
                logger.warning("Circular reference detected in tree structure")
                return
            
            visited.add(id(node))
            
            try:
                self._validate_node_attributes(node)
                
                if hasattr(node, 'is_terminal') and node.is_terminal:
                    if hasattr(node, 'samples') and node.samples > 0:
                        terminal_nodes.append(node)
                    else:
                        logger.warning(f"Terminal node {getattr(node, 'node_id', 'unknown')} has zero samples - excluding")
                elif hasattr(node, 'children') and node.children:
                    for child in node.children:
                        if child:  # Ensure child is not None
                            if hasattr(child, 'parent') and child.parent != node:
                                logger.warning(f"Inconsistent parent-child relationship for node {getattr(child, 'node_id', 'unknown')}")
                            traverse(child, depth + 1)
            except Exception as e:
                logger.warning(f"Error traversing node {getattr(node, 'node_id', 'unknown')}: {e}")
        
        traverse(root_node)
        return terminal_nodes
    
    def _validate_node_attributes(self, node: TreeNode) -> None:
        """Validate that node has required attributes"""
        try:
            if not hasattr(node, 'node_id'):
                logger.debug(f"Node missing node_id attribute")
            
            if not hasattr(node, 'is_terminal'):
                logger.debug(f"Node {getattr(node, 'node_id', 'unknown')} missing is_terminal attribute")
            
            if hasattr(node, 'is_terminal') and not node.is_terminal:
                required_attrs = ['split_feature', 'split_value', 'split_type']
                missing_attrs = []
                for attr in required_attrs:
                    if not hasattr(node, attr) or getattr(node, attr) is None:
                        missing_attrs.append(attr)
                
                if missing_attrs:
                    logger.warning(f"Non-terminal node {getattr(node, 'node_id', 'unknown')} "
                                 f"missing split attributes: {missing_attrs}")
            
        except Exception as e:
            logger.debug(f"Error validating node attributes: {e}")
    
    def _validate_calculation_results(self, node_data: List[Dict], overall_stats: Dict[str, Any]) -> None:
        """Validate calculation results for edge cases"""
        try:
            if not node_data:
                return
            
            zero_sample_nodes = [node for node in node_data if node.get('node_size', 0) == 0]
            if zero_sample_nodes:
                logger.warning(f"Found {len(zero_sample_nodes)} nodes with zero samples")
            
            target_rates = [node.get('target_rate', 0) for node in node_data]
            unique_rates = set(target_rates)
            if len(unique_rates) == 1:
                logger.warning("All nodes have identical target rates - lift calculations may not be meaningful")
            
            lifts = [node.get('lift', 0) for node in node_data if node.get('lift', 0) != 0]
            if lifts:
                max_lift = max(lifts)
                min_lift = min(lifts)
                if max_lift > 1000:  # 10x lift seems extreme
                    logger.warning(f"Extremely high lift values detected (max: {max_lift:.1f})")
                if min_lift < 10:  # Less than 10% of baseline
                    logger.warning(f"Extremely low lift values detected (min: {min_lift:.1f})")
            
            total_node_samples = sum(node.get('node_size', 0) for node in node_data)
            expected_total = overall_stats.get('total_size', 0)
            if abs(total_node_samples - expected_total) > expected_total * 0.01:  # 1% tolerance
                logger.warning(f"Sample count mismatch: nodes={total_node_samples}, "
                             f"expected={expected_total}")
            
        except Exception as e:
            logger.debug(f"Error validating calculation results: {e}")
    
    def _calculate_lift_metrics(self, node_stats: Dict[str, Any], overall_stats: Dict[str, Any]) -> Dict[str, Any]:
        """
        Calculate lift metrics for a node as specified in implementation guide
        
        Args:
            node_stats: Dictionary containing node statistics including target_rate
            overall_stats: Dictionary containing overall dataset statistics
            
        Returns:
            Dictionary containing lift metrics
        """
        try:
            node_target_rate = node_stats.get('target_rate', 0.0)
            overall_target_rate = overall_stats.get('target_rate', 0.0)
            
            lift = self._calculate_lift(node_target_rate, overall_target_rate)
            
            return {
                'lift': round(lift, 1),
                'lift_ratio': round(node_target_rate / overall_target_rate, 3) if overall_target_rate > 0 else 0.0,
                'performance_vs_baseline': 'Above Average' if lift > 100 else 'Below Average' if lift < 100 else 'Average'
            }
            
        except Exception as e:
            logger.warning(f"Error calculating lift metrics: {e}")
            return {
                'lift': 0.0,
                'lift_ratio': 0.0,
                'performance_vs_baseline': 'Unknown'
            }
    
    def _get_tree_level_statistics(self, model) -> Dict[str, Any]:
        """
        Get overall dataset statistics from ROOT NODE STORED VALUES (CORRECT APPROACH)
        Uses the exact same stored data that tree visualization uses
        """
        try:
            root_node = model.root
            if not root_node:
                logger.error("Model has no root node")
                return {'total_size': 0, 'target_count': 0, 'target_rate': 0.0}
            
            total_size = getattr(root_node, 'samples', 0)
            class_counts = getattr(root_node, 'class_counts', {})
            
            logger.debug(f"Root node stored values: samples={total_size}, class_counts={class_counts}")
            
            positive_class = None
            if class_counts:
                if 1 in class_counts:
                    positive_class = 1
                elif True in class_counts:
                    positive_class = True
                else:
                    numeric_classes = [k for k in class_counts.keys() if isinstance(k, (int, float))]
                    if numeric_classes:
                        positive_class = max(numeric_classes)
                    else:
                        positive_class = list(class_counts.keys())[0] if class_counts else None
            
            target_count = class_counts.get(positive_class, 0) if positive_class is not None else 0
            
            target_rate = self._safe_divide(target_count * 100, total_size, 0.0)
            
            logger.debug(f"Tree-level stats (STORED): total_size={total_size}, target_count={target_count}, "
                        f"target_rate={target_rate:.2f}%, positive_class={positive_class}")
            
            return {
                'total_size': total_size,
                'target_count': target_count,
                'target_rate': target_rate,
                'positive_class': positive_class
            }
            
        except Exception as e:
            logger.error(f"Error getting tree-level statistics: {e}")
            return {'total_size': 0, 'target_count': 0, 'target_rate': 0.0}
    
    def _determine_positive_class_from_tree(self, class_counts: Dict) -> Any:
        """Determine positive class from tree class counts"""
        if len(class_counts) == 2:
            for class_val in class_counts.keys():
                if class_val in [1, True, 'Yes', 'yes', 'Y', 'y', 'True', 'true', 'positive', 'Positive']:
                    return class_val
            return sorted(class_counts.keys())[-1]
        else:
            return max(class_counts.keys(), key=lambda k: class_counts[k])
    
    def _determine_positive_class(self, target_values: pd.Series) -> Any:
        """Determine positive class for binary classification"""
        unique_values = target_values.unique()
        
        if len(unique_values) == 2:
            for val in unique_values:
                if val in [1, True, 'Yes', 'yes', 'Y', 'y', 'True', 'true', 'positive', 'Positive']:
                    self._positive_class = val
                    return val
            
            sorted_vals = sorted(unique_values)
            self._positive_class = sorted_vals[-1]  # Usually 1 is after 0, True after False
            return self._positive_class
        
        most_common = target_values.value_counts().index[0]
        self._positive_class = most_common
        return most_common
    
    def _calculate_node_statistics(self, node: TreeNode, overall_stats: Dict[str, Any], node_number: int = 0) -> Optional[Dict[str, Any]]:
        """
        Calculate comprehensive statistics for a single node using STORED TREE NODE VALUES (CORRECT APPROACH)
        Uses the exact same stored data that tree visualization uses
        
        Args:
            node: TreeNode to analyze  
            overall_stats: Overall dataset statistics
            node_number: Sequential node number for display
            
        Returns:
            Dictionary with node statistics or None if calculation failed
        """
        try:
            node_logic = self._format_node_logic(node)
            
            node_id = self._generate_node_id(node)
            
            node_size = getattr(node, 'samples', 0)
            class_counts = getattr(node, 'class_counts', {})
            
            logger.debug(f"Node {node.node_id} stored values: samples={node_size}, class_counts={class_counts}")
            
            if node_size == 0:
                logger.warning(f"Node {node.node_id} has no samples stored")
                return None
                
            if not class_counts:
                logger.warning(f"Node {node.node_id} has no class_counts stored")
                return None
            
            positive_class = overall_stats.get('positive_class')
            if positive_class is not None:
                target_count = class_counts.get(positive_class, 0)
            else:
                target_count = sum(count for class_val, count in class_counts.items() 
                                 if class_val in [1, True, 'Yes', 'yes', 'Y', 'y'])
            
            target_rate = self._safe_divide(target_count * 100, node_size, 0.0)
            
            lift_metrics = self._calculate_lift_metrics(
                {'target_rate': target_rate}, 
                overall_stats
            )
            lift = lift_metrics.get('lift', 0.0)
            
            logger.debug(f"Node {node.node_id} final stats: size={node_size}, target_count={target_count}, rate={target_rate:.2f}%")
            
            return {
                'node_no': node_number,
                'node_logic': node_logic,
                'node_id': node_id,
                'node_size': node_size,
                'target_count': target_count,
                'target_rate': round(target_rate, 2),
                'lift': round(lift, 1),
                'tree_node': node  # Keep reference for sorting
            }
            
        except Exception as e:
            logger.warning(f"Error calculating statistics for node {getattr(node, 'node_id', 'unknown')}: {e}")
            return None
    
    def _format_node_logic(self, node: TreeNode) -> str:
        """
        Format node decision path using pre-computed split rules from tree
        
        Args:
            node: TreeNode to format
            
        Returns:
            Formatted logic string using actual tree split data
        """
        try:
            if hasattr(node, 'decision_path') and node.decision_path:
                return node.decision_path
            
            conditions = []
            current = node
            
            while current and current.parent:
                parent = current.parent
                
                if hasattr(parent, 'split_rule') and parent.split_rule:
                    child_condition = self._get_child_specific_condition(parent, current)
                    if child_condition:
                        conditions.append(child_condition)
                else:
                    condition = self._build_condition(parent, current)
                    if condition:
                        conditions.append(condition)
                
                current = parent
            
            conditions.reverse()
            
            if conditions:
                return " AND ".join(conditions)
            else:
                return "Root Node"
                
        except Exception as e:
            logger.warning(f"Could not format logic for node {getattr(node, 'node_id', 'unknown')}: {e}")
            return "Terminal Node"
    
    def _get_child_specific_condition(self, parent_node: TreeNode, child_node: TreeNode) -> Optional[str]:
        """Extract the specific condition for this child from the parent's split rule"""
        try:
            try:
                child_index = parent_node.children.index(child_node)
            except (ValueError, AttributeError):
                return None
            
            split_type = getattr(parent_node, 'split_type', 'numeric')
            feature = getattr(parent_node, 'split_feature', None)
            
            if not feature:
                return None
            
            if split_type == 'categorical':
                categories = getattr(parent_node, 'split_categories', {})
                if categories:
                    child_categories = [cat for cat, idx in categories.items() if idx == child_index]
                    if child_categories:
                        if len(child_categories) == 1:
                            return f"[{feature}] = '{child_categories[0]}'"
                        else:
                            cat_str = "', '".join(child_categories)
                            return f"[{feature}] IN ['{cat_str}']"
            
            elif split_type == 'numeric':
                split_value = getattr(parent_node, 'split_value', None)
                if split_value is not None:
                    if child_index == 0:  # Left child
                        return f"[{feature}] <= {split_value}"
                    else:  # Right child
                        return f"[{feature}] > {split_value}"
            
            split_rule = getattr(parent_node, 'split_rule', None)
            if split_rule:
                return f"[{feature}] ({split_rule})"
                
            return None
            
        except Exception as e:
            logger.debug(f"Error getting child-specific condition: {e}")
            return None
    
    def _build_condition(self, parent_node: TreeNode, child_node: TreeNode) -> Optional[str]:
        """Build condition string for parent->child relationship using the same logic as tree visualizer"""
        try:
            feature = getattr(parent_node, 'split_feature', None)
            if not feature:
                logger.debug(f"Parent node {getattr(parent_node, 'node_id', 'unknown')} has no split_feature")
                return None
            
            split_type = getattr(parent_node, 'split_type', 'numeric')
            logger.debug(f"Building condition for feature '{feature}' with split_type '{split_type}'")
            
            try:
                child_index = parent_node.children.index(child_node)
            except (ValueError, AttributeError):
                logger.debug(f"Could not determine child index for node relationship")
                return f"[{feature}] (index not found)"
            
            if split_type == 'numeric':
                threshold = getattr(parent_node, 'split_value', 0)
                if child_index == 0:  # Left child
                    return f"[{feature}] <= {threshold:.3f}"
                else:  # Right child
                    return f"[{feature}] > {threshold:.3f}"
                        
            elif split_type == 'numeric_multi_bin':
                if hasattr(parent_node, 'split_thresholds') and parent_node.split_thresholds:
                    thresholds = parent_node.split_thresholds
                    if child_index == 0:
                        return f"[{feature}] <= {thresholds[0]}"
                    elif child_index == len(thresholds):
                        return f"[{feature}] > {thresholds[-1]}"
                    else:
                        return f"[{feature}] > {thresholds[child_index-1]} AND [{feature}] <= {thresholds[child_index]}"
                else:
                    return f"[{feature}] (bin {child_index})"
                        
            elif split_type == 'categorical':
                categories = getattr(parent_node, 'split_categories', {})
                if categories:
                    child_categories = [cat for cat, idx in categories.items() if idx == child_index]
                    if child_categories:
                        if len(child_categories) == 1:
                            return f"[{feature}] = '{child_categories[0]}'"
                        else:
                            cat_str = "', '".join(child_categories)
                            return f"[{feature}] IN ['{cat_str}']"
                    else:
                        return f"[{feature}] (no categories for child {child_index})"
                else:
                    return f"[{feature}] (categorical)"
            
            elif split_type == 'categorical_multi_bin':
                if hasattr(parent_node, 'split_bin_categories') and parent_node.split_bin_categories:
                    bin_categories = parent_node.split_bin_categories
                    if child_index < len(bin_categories):
                        categories = bin_categories[child_index]
                        if len(categories) <= 3:
                            cat_str = ", ".join(str(cat) for cat in categories)
                            return f"[{feature}] IN [{cat_str}]"
                        else:
                            return f"[{feature}] IN [{len(categories)} categories]"
                    else:
                        return f"[{feature}] (bin {child_index})"
                else:
                    return f"[{feature}] (categorical bin {child_index})"
            
            elif split_type == 'missing':
                if child_index == 0:  # Missing branch
                    return f"[{feature}] IS NULL"
                else:  # Not missing branch
                    return f"[{feature}] IS NOT NULL"
            
            else:
                split_value = getattr(parent_node, 'split_value', 'N/A')
                return f"[{feature}] = {split_value} ({split_type})"
        
        except Exception as e:
            logger.debug(f"Could not build condition: {e}")
            feature = getattr(parent_node, 'split_feature', 'Unknown')
            return f"[{feature}] (condition parsing failed)"
            
        return None
    
    def _create_node_number_mapping(self, root_node: TreeNode) -> Dict[str, int]:
        """
        Create node number mapping using EXACT SAME algorithm as tree visualizer
        This ensures perfect consistency between tree visualization and node report numbering
        
        Args:
            root_node: Root node of the decision tree
            
        Returns:
            Dictionary mapping node_id to visualization node number
        """
        try:
            all_nodes = root_node.get_subtree_nodes()  # Gets ALL nodes (internal + terminal)
            
            if not all_nodes:
                logger.warning("No nodes found in tree for numbering")
                return {}
            
            node_number_mapping = {}
            for i, node in enumerate(all_nodes):
                node_number_mapping[node.node_id] = i + 1  # Sequential: 1,2,3,4,5,6,7,8,9,10,11,12,13,14,15
            
            logger.debug(f"Created node number mapping using tree visualizer algorithm: {len(node_number_mapping)} nodes")
            logger.debug(f"Terminal nodes will have numbers: {[node_number_mapping[n.node_id] for n in all_nodes if getattr(n, 'is_terminal', False)]}")
            
            return node_number_mapping
            
        except Exception as e:
            logger.error(f"Error creating node number mapping: {e}")
            return {}
    
    
    def _build_alternative_condition(self, parent_node: TreeNode, child_node: TreeNode) -> Optional[str]:
        """Alternative method to build condition when primary method fails"""
        try:
            split_rule = getattr(parent_node, 'split_rule', None)
            if split_rule:
                return split_rule
            
            feature = getattr(parent_node, 'split_feature', None)
            if feature:
                split_value = getattr(parent_node, 'split_value', None)
                if split_value is not None:
                    if hasattr(parent_node, 'children') and parent_node.children:
                        try:
                            child_index = parent_node.children.index(child_node)
                            if child_index == 0:
                                return f"[{feature}] <= {split_value}"
                            else:
                                return f"[{feature}] > {split_value}"
                        except ValueError:
                            return f"[{feature}] ~ {split_value}"
                    else:
                        return f"[{feature}] ~ {split_value}"
                else:
                    return f"[{feature}] (no threshold)"
            
            return None
            
        except Exception as e:
            logger.debug(f"Alternative condition building failed: {e}")
            return None
    
    def _generate_node_id(self, node: TreeNode) -> str:
        """Generate hierarchical node ID reflecting tree structure"""
        try:
            path_components = []
            current = node
            
            while current and current.parent:
                parent = current.parent
                if hasattr(parent, 'children') and parent.children:
                    child_index = parent.children.index(current)
                    path_components.append(str(child_index + 1))  # 1-based indexing
                current = parent
            
            path_components.reverse()
            
            if path_components:
                return "/".join(["x"] + path_components)  # Start with 'x' like in the example
            else:
                return "x0"  # Root node
                
        except Exception as e:
            logger.debug(f"Could not generate node ID: {e}")
            return f"node_{getattr(node, 'node_id', 'unknown')}"
    
    def _calculate_cumulative_metrics(self, df: pd.DataFrame) -> pd.DataFrame:
        """Calculate cumulative metrics for the node report"""
        try:
            df = df.sort_values('node_size', ascending=False).copy()
            
            df['cumulative_size'] = df['node_size'].cumsum()
            df['cumulative_target_count'] = df['target_count'].cumsum()
            
            total_size = df['node_size'].sum()
            if total_size > 0:
                df['cumulative_pct'] = (df['cumulative_size'] / total_size * 100).round(2)
                df['cumulative_target_rate'] = (
                    df['cumulative_target_count'] / df['cumulative_size'] * 100
                ).fillna(0).round(2)
            else:
                df['cumulative_pct'] = 0.0
                df['cumulative_target_rate'] = 0.0
            
            if 'tree_node' in df.columns:
                df = df.drop('tree_node', axis=1)
            
            column_order = [
                'node_no', 'node_logic', 'node_id', 'node_size', 'cumulative_size', 
                'cumulative_pct', 'target_count', 'cumulative_target_count',
                'target_rate', 'cumulative_target_rate', 'lift'
            ]
            
            existing_columns = [col for col in column_order if col in df.columns]
            df = df[existing_columns]
            
            return df
            
        except Exception as e:
            logger.error(f"Error calculating cumulative metrics: {e}")
            return df
    
    def _calculate_lift(self, node_target_rate: float, overall_target_rate: float) -> float:
        """Calculate lift as node rate vs overall rate"""
        try:
            if overall_target_rate == 0:
                return 0.0  # Cannot calculate lift with zero baseline
            
            lift = (node_target_rate / overall_target_rate) * 100
            return lift if not pd.isna(lift) else 0.0
            
        except (ZeroDivisionError, TypeError):
            return 0.0
    
    def _safe_divide(self, numerator: float, denominator: float, default: float = 0.0) -> float:
        """Safely divide two numbers with fallback for edge cases"""
        try:
            if denominator == 0 or pd.isna(denominator):
                return default
            result = numerator / denominator
            return result if not pd.isna(result) else default
        except (ZeroDivisionError, TypeError):
            return default
    
    def export_to_excel(self, df: pd.DataFrame, file_path: str, 
                       sheet_name: str = 'Node Report', timeout_seconds: int = 300) -> Tuple[bool, str]:
        """
        Export node report DataFrame to Excel with professional formatting and timeout handling
        
        Args:
            df: DataFrame to export
            file_path: Path to save Excel file
            sheet_name: Name of Excel sheet
            timeout_seconds: Maximum time to wait for export (default: 5 minutes)
            
        Returns:
            Tuple of (success: bool, message: str)
        """
        import signal
        
        def timeout_handler(signum, frame):
            raise TimeoutError("Excel export timed out")
        
        try:
            try:
                import openpyxl
            except ImportError:
                return False, "openpyxl library is required for Excel export. Install with: pip install openpyxl"
            
            Path(file_path).parent.mkdir(parents=True, exist_ok=True)
            
            if len(df) > 1000:  # Only use timeout for large datasets
                signal.signal(signal.SIGALRM, timeout_handler)
                signal.alarm(timeout_seconds)
            
            try:
                with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
                    df.to_excel(writer, sheet_name=sheet_name, index=False)
                    
                    self._format_excel_sheet(writer, sheet_name, df)
                
                if len(df) > 1000:
                    signal.alarm(0)
                
                if not Path(file_path).exists():
                    return False, "Excel file was not created successfully"
                
                file_size = Path(file_path).stat().st_size
                if file_size == 0:
                    return False, "Excel file was created but is empty"
                
                return True, f"Node report exported successfully to {file_path} ({file_size:,} bytes)"
                
            except TimeoutError:
                return False, f"Export timed out after {timeout_seconds} seconds. Dataset too large."
            finally:
                if len(df) > 1000:
                    signal.alarm(0)
            
        except PermissionError:
            return False, "Permission denied. File may be open in another application."
        except FileNotFoundError:
            return False, "Invalid file path or directory does not exist."
        except Exception as e:
            return False, f"Export failed: {str(e)}"
    
    def _format_excel_sheet(self, writer, sheet_name: str, df: pd.DataFrame) -> None:
        """Apply professional formatting to Excel sheet"""
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
            
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            border = Border(left=Side(border_style="thin"), right=Side(border_style="thin"),
                          top=Side(border_style="thin"), bottom=Side(border_style="thin"))
            
            for col_num, value in enumerate(df.columns.values, 1):
                cell = worksheet.cell(row=1, column=col_num)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = border
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            worksheet.column_dimensions['A'].width = 8   # Node No column (narrow)
            worksheet.column_dimensions['B'].width = 50  # Node Logic column (wide)
            worksheet.column_dimensions['C'].width = 15  # Node ID
            for col_letter in ['D', 'E', 'F', 'G', 'H', 'I', 'J', 'K']:
                worksheet.column_dimensions[col_letter].width = 12
            
            for row_num in range(2, len(df) + 2):  # Start from row 2 (after header)
                for col_num, column_name in enumerate(df.columns, 1):
                    cell = worksheet.cell(row=row_num, column=col_num)
                    
                    if column_name in ['node_no', 'node_size', 'cumulative_size', 'target_count', 
                                     'cumulative_target_count', 'target_rate', 
                                     'cumulative_target_rate', 'cumulative_pct', 'lift']:
                        cell.alignment = Alignment(horizontal="right")
                    
                    if column_name == 'node_logic':
                        cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            worksheet.freeze_panes = "A2"
            
        except Exception as e:
            logger.warning(f"Could not apply Excel formatting: {e}")